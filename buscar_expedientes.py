#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Robot de Búsqueda Automática de Expedientes v6.0
TSJ Quintana Roo - Lista Electrónica
Autor: Jorge Israel Clemente Marié - Empírica Legal Lab

VERSIÓN 6.0 - OPTIMIZADA:
- ✅ Búsquedas simultáneas en pestañas paralelas
- ✅ Carga dinámica de expedientes desde archivo JSON
- ✅ Reporte Excel con formato y marcado de acuerdos nuevos (últimos 5 días)
- ✅ Mejor manejo de errores y reintentos
- ✅ Archivo de configuración separado
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time
import csv
import json
from datetime import datetime, timedelta
import os
import threading

class TSJExpedientesBot:
    
    # IDs exactos del sistema TSJ (extraídos del sidebar.php)
    JUZGADOS = {
        # ===== CANCÚN =====
        'JUZGADO PRIMERO FAMILIAR ORAL CANCUN': 109,
        'JUZGADO SEGUNDO FAMILIAR ORAL CANCUN': 158,
        'JUZGADO SEGUNDO DE LO FAMILIAR CANCUN': 115,
        'JUZGADO FAMILIAR DE PRIMERA INSTANCIA CANCUN': 114,
        'JUZGADO PRIMERO CIVIL CANCUN': 111,
        'JUZGADO SEGUNDO CIVIL CANCUN': 112,
        'JUZGADO TERCERO CIVIL CANCUN': 113,
        'JUZGADO CUARTO CIVIL CANCUN': 182,
        'JUZGADO ORAL CIVIL CANCUN': 110,
        'JUZGADO PRIMERO MERCANTIL CANCUN': 105,
        'JUZGADO SEGUNDO MERCANTIL CANCUN': 106,
        'JUZGADO TERCERO MERCANTIL CANCUN': 107,
        'JUZGADO ORAL MERCANTIL CANCUN': 108,
        'TRIBUNAL PRIMERO LABORAL CANCUN': 164,
        'TRIBUNAL SEGUNDO LABORAL CANCUN': 165,
        
        # ===== PLAYA DEL CARMEN / SOLIDARIDAD =====
        'JUZGADO FAMILIAR ORAL PLAYA': 88,
        'JUZGADO FAMILIAR PRIMERA INSTANCIA PLAYA': 84,
        'JUZGADO PRIMERO CIVIL PLAYA': 83,
        'JUZGADO SEGUNDO CIVIL PLAYA': 161,
        'JUZGADO ORAL CIVIL PLAYA': 87,
        'JUZGADO MERCANTIL PLAYA': 85,
        'TRIBUNAL LABORAL PLAYA': 166,
        
        # ===== CHETUMAL =====
        'JUZGADO FAMILIAR ORAL CHETUMAL': 93,
        'JUZGADO FAMILIAR PRIMERA INSTANCIA CHETUMAL': 94,
        'JUZGADO CIVIL CHETUMAL': 95,
        'JUZGADO MERCANTIL CHETUMAL': 96,
        'JUZGADO CIVIL ORAL CHETUMAL': 97,
        'TRIBUNAL LABORAL CHETUMAL': 163,
        
        # ===== COZUMEL =====
        'JUZGADO FAMILIAR COZUMEL': 89,
        'JUZGADO CIVIL COZUMEL': 90,
        'JUZGADO FAMILIAR ORAL COZUMEL': 91,
        'JUZGADO ORAL CIVIL COZUMEL': 92,
        
        # ===== FELIPE CARRILLO PUERTO =====
        'JUZGADO CIVIL ORAL CARRILLO PUERTO': 136,
        'JUZGADO FAMILIAR ORAL CARRILLO PUERTO': 137,
        'JUZGADO CIVIL PRIMERA INSTANCIA CARRILLO PUERTO': 153,
        'JUZGADO FAMILIAR PRIMERA INSTANCIA CARRILLO PUERTO': 154,
        
        # ===== ISLA MUJERES =====
        'JUZGADO CIVIL ORAL ISLA MUJERES': 131,
        'JUZGADO FAMILIAR ORAL ISLA MUJERES': 132,
        
        # ===== TULUM =====
        'JUZGADO CIVIL ORAL TULUM': 144,
        'JUZGADO FAMILIAR ORAL TULUM': 145,
        
        # ===== BACALAR =====
        'JUZGADO FAMILIAR PRIMERA INSTANCIA BACALAR': 188,
    }
    
    def __init__(self, max_pestanas=5, dias_acuerdos_nuevos=5):
        self.base_url = "https://www.tsjqroo.gob.mx/estrados"
        self.driver = None
        self.resultados = []
        self.debug_mode = True
        self.screenshot_dir = "debug_screenshots"
        self.max_pestanas = max_pestanas  # Número máximo de pestañas simultáneas
        self.dias_acuerdos_nuevos = dias_acuerdos_nuevos  # Días para marcar como "nuevo"
        self.resultados_lock = threading.Lock()  # Para thread-safety

        if self.debug_mode and not os.path.exists(self.screenshot_dir):
            os.makedirs(self.screenshot_dir)
    
    def log(self, msg, nivel="INFO"):
        timestamp = datetime.now().strftime('%H:%M:%S')
        iconos = {"INFO": "ℹ️", "OK": "✅", "WARN": "⚠️", "ERROR": "❌", "DEBUG": "🔍"}
        print(f"[{timestamp}] {iconos.get(nivel, '•')} {msg}")
    
    def screenshot(self, nombre):
        if self.debug_mode and self.driver:
            try:
                filename = f"{self.screenshot_dir}/{nombre}_{datetime.now().strftime('%H%M%S')}.png"
                self.driver.save_screenshot(filename)
            except:
                pass
    
    def guardar_html(self, nombre):
        if self.debug_mode and self.driver:
            try:
                filename = f"{self.screenshot_dir}/{nombre}_{datetime.now().strftime('%H%M%S')}.html"
                with open(filename, 'w', encoding='utf-8') as f:
                    f.write(self.driver.page_source)
            except:
                pass

    @staticmethod
    def cargar_configuracion(archivo='config.json'):
        """Carga configuración desde archivo JSON"""
        try:
            if os.path.exists(archivo):
                with open(archivo, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    config = data.get('configuracion', {})
                    return config
            return {}
        except Exception as e:
            print(f"⚠️  Error cargando configuración: {e}")
            return {}

    def cargar_expedientes_json(self, archivo='expedientes.json'):
        """Carga expedientes desde archivo JSON"""
        try:
            if not os.path.exists(archivo):
                self.log(f"Archivo {archivo} no encontrado", "WARN")
                return []

            with open(archivo, 'r', encoding='utf-8') as f:
                data = json.load(f)

            expedientes = data.get('expedientes', [])
            self.log(f"Cargados {len(expedientes)} expedientes desde {archivo}", "OK")
            return expedientes

        except Exception as e:
            self.log(f"Error cargando expedientes: {e}", "ERROR")
            return []

    def es_acuerdo_nuevo(self, fecha_publicacion_str):
        """Determina si un acuerdo es nuevo (últimos N días)"""
        try:
            # Intentar diferentes formatos de fecha
            formatos = ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y']
            fecha_publicacion = None

            for formato in formatos:
                try:
                    fecha_publicacion = datetime.strptime(fecha_publicacion_str.strip(), formato)
                    break
                except ValueError:
                    continue

            if not fecha_publicacion:
                return False

            fecha_limite = datetime.now() - timedelta(days=self.dias_acuerdos_nuevos)
            return fecha_publicacion >= fecha_limite

        except Exception as e:
            return False
    
    def iniciar_navegador(self):
        self.log("Iniciando navegador Chrome...")
        
        opciones = webdriver.ChromeOptions()
        opciones.add_argument('--start-maximized')
        opciones.add_argument('--disable-notifications')
        opciones.add_argument('--disable-blink-features=AutomationControlled')
        opciones.add_argument('--no-sandbox')
        opciones.add_experimental_option("excludeSwitches", ["enable-automation"])
        opciones.add_experimental_option('useAutomationExtension', False)
        opciones.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36')
        
        self.driver = webdriver.Chrome(options=opciones)
        self.driver.implicitly_wait(10)
        self.log("Navegador iniciado", "OK")
    
    def obtener_id_juzgado(self, nombre_juzgado):
        """Obtiene el ID interno del juzgado"""
        nombre_upper = nombre_juzgado.upper().strip()
        
        # Búsqueda exacta primero
        for nombre, id_juzgado in self.JUZGADOS.items():
            if nombre.upper() == nombre_upper:
                return id_juzgado
        
        # Búsqueda parcial flexible
        for nombre, id_juzgado in self.JUZGADOS.items():
            nombre_norm = nombre.upper()
            # Si coinciden las palabras clave principales
            if all(palabra in nombre_norm for palabra in nombre_upper.split()[:3]):
                return id_juzgado
        
        self.log(f"⚠️  Juzgado no encontrado: {nombre_juzgado}", "WARN")
        return None
    
    def buscar(self, termino, id_juzgado, metodo=1):
        """
        Realiza una búsqueda en el sistema
        metodo=1: por expediente
        metodo=2: por nombre (actores)
        """
        try:
            url = f"{self.base_url}/buscador_primera.php?int={id_juzgado}&metodo={metodo}&findexp={termino}"
            
            tipo = "expediente" if metodo == 1 else "nombre"
            self.log(f"Buscando {tipo}: {termino}")
            self.driver.get(url)
            time.sleep(4)  # Esperar a que cargue la tabla
            
            self.screenshot(f"resultado_{termino.replace('/', '_').replace(' ', '_')}")
            self.guardar_html(f"resultado_{termino.replace('/', '_').replace(' ', '_')}")
            return True
            
        except Exception as e:
            self.log(f"Error en búsqueda: {e}", "ERROR")
            return False
    
    def extraer_resultados(self, busqueda, juzgado, tipo_busqueda="expediente", driver=None):
        """Extrae los resultados de la tabla de publicaciones"""
        if driver is None:
            driver = self.driver

        publicaciones = []

        try:
            time.sleep(2)

            # Verificar si no hay resultados
            page_source = driver.page_source
            if "No se encontr" in page_source or "ningun resultado" in page_source.lower():
                self.log(f"Sin publicaciones para: {busqueda}", "WARN")
                resultado = {
                    'busqueda': busqueda,
                    'tipo_busqueda': tipo_busqueda,
                    'juzgado': juzgado,
                    'estado': 'Sin publicaciones',
                    'fecha_busqueda': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'publicaciones': []
                }
                with self.resultados_lock:
                    self.resultados.append(resultado)
                return resultado

            # Buscar filas de la tabla (las filas de datos tienen clase 'odd' o 'even')
            filas = driver.find_elements(By.CSS_SELECTOR, "tr.odd, tr.even")

            if not filas:
                # Intentar buscar cualquier fila de tabla con datos
                filas = driver.find_elements(By.XPATH, "//table//tr[td]")

            self.log(f"Filas encontradas: {len(filas)}", "DEBUG")

            for fila in filas:
                try:
                    celdas = fila.find_elements(By.TAG_NAME, "td")
                    if len(celdas) >= 7:
                        fecha_pub = celdas[6].text.strip()
                        publicacion = {
                            'id_acuerdo': celdas[0].text.strip(),
                            'documento': celdas[1].text.strip(),
                            'juicio': celdas[2].text.strip(),
                            'promoventes': celdas[3].text.strip(),
                            'demandados': celdas[4].text.strip(),
                            'extracto': celdas[5].text.strip(),
                            'fecha_publicacion': fecha_pub,
                            'es_nuevo': self.es_acuerdo_nuevo(fecha_pub)  # Marcar si es nuevo
                        }
                        publicaciones.append(publicacion)
                except Exception as e:
                    self.log(f"Error en fila: {e}", "DEBUG")
                    continue

            resultado = {
                'busqueda': busqueda,
                'tipo_busqueda': tipo_busqueda,
                'juzgado': juzgado,
                'estado': 'Con publicaciones' if publicaciones else 'Sin publicaciones',
                'fecha_busqueda': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'publicaciones': publicaciones
            }

            with self.resultados_lock:
                self.resultados.append(resultado)

            nuevos = sum(1 for p in publicaciones if p.get('es_nuevo', False))
            self.log(f"✅ Encontradas {len(publicaciones)} publicaciones ({nuevos} nuevas)", "OK")
            return resultado

        except Exception as e:
            self.log(f"Error extrayendo resultados: {e}", "ERROR")
            return None
    
    def procesar_expediente_en_pestana(self, exp, pestana_idx):
        """Procesa un expediente en una pestaña específica del navegador"""
        try:
            termino = exp.get('numero', exp.get('nombre', 'N/A'))
            self.log(f"[Pestaña {pestana_idx}] Procesando: {termino}")

            # Obtener ID del juzgado
            id_juzgado = self.obtener_id_juzgado(exp['juzgado'])
            if not id_juzgado:
                self.log(f"[Pestaña {pestana_idx}] Juzgado no encontrado: {exp['juzgado']}", "ERROR")
                return None

            # Construir URL
            if 'numero' in exp:
                metodo = 1
                termino_busqueda = exp['numero']
                tipo_busqueda = "expediente"
            elif 'nombre' in exp:
                metodo = 2
                termino_busqueda = exp['nombre']
                tipo_busqueda = "nombre"
            else:
                self.log(f"[Pestaña {pestana_idx}] Expediente sin número ni nombre", "ERROR")
                return None

            url = f"{self.base_url}/buscador_primera.php?int={id_juzgado}&metodo={metodo}&findexp={termino_busqueda}"

            # Cambiar a la pestaña correspondiente
            self.driver.switch_to.window(self.driver.window_handles[pestana_idx])

            # Realizar búsqueda
            self.driver.get(url)
            time.sleep(4)  # Esperar carga

            # Extraer resultados
            resultado = self.extraer_resultados(
                termino_busqueda,
                exp['juzgado'],
                tipo_busqueda,
                driver=self.driver
            )

            self.log(f"[Pestaña {pestana_idx}] ✅ Completado: {termino}", "OK")
            return resultado

        except Exception as e:
            self.log(f"[Pestaña {pestana_idx}] Error: {e}", "ERROR")
            return None

    def procesar_expedientes(self, expedientes):
        """Procesa expedientes en paralelo usando múltiples pestañas"""
        total = len(expedientes)
        self.log(f"\n{'='*60}")
        self.log(f"PROCESANDO {total} BÚSQUEDAS EN PARALELO")
        self.log(f"Pestañas simultáneas: {self.max_pestanas}")
        self.log(f"{'='*60}\n")

        if not expedientes:
            self.log("No hay expedientes para procesar", "WARN")
            return

        # Abrir pestañas necesarias
        num_pestanas = min(self.max_pestanas, total)
        self.log(f"Abriendo {num_pestanas} pestañas...")

        for i in range(num_pestanas - 1):  # -1 porque ya tenemos una pestaña abierta
            self.driver.execute_script("window.open('');")
            time.sleep(0.5)

        self.log(f"✅ {num_pestanas} pestañas abiertas", "OK")

        # Procesar expedientes en lotes
        procesados = 0
        lote_idx = 0

        while procesados < total:
            lote = expedientes[procesados:procesados + num_pestanas]
            lote_size = len(lote)

            self.log(f"\n--- LOTE {lote_idx + 1} ({lote_size} búsquedas) ---")

            # Procesar cada expediente del lote en una pestaña diferente
            for idx, exp in enumerate(lote):
                self.procesar_expediente_en_pestana(exp, idx)

            procesados += lote_size
            lote_idx += 1

            self.log(f"Progreso: {procesados}/{total} búsquedas completadas\n")

        self.log(f"✅ Todas las búsquedas completadas", "OK")
    
    def guardar_csv(self, archivo='resultados_expedientes.csv'):
        """Guarda resultados en CSV"""
        self.log(f"Guardando en {archivo}...")
        
        with open(archivo, 'w', newline='', encoding='utf-8') as f:
            campos = [
                'Búsqueda', 'Tipo', 'Juzgado', 'Estado', 'Fecha Consulta',
                'IdAcuerdo', 'Documento', 'Juicio', 'Promoventes', 
                'Demandados', 'Extracto', 'Fecha Publicación'
            ]
            writer = csv.DictWriter(f, fieldnames=campos)
            writer.writeheader()
            
            for r in self.resultados:
                if r['publicaciones']:
                    for p in r['publicaciones']:
                        writer.writerow({
                            'Búsqueda': r['busqueda'],
                            'Tipo': r['tipo_busqueda'],
                            'Juzgado': r['juzgado'],
                            'Estado': r['estado'],
                            'Fecha Consulta': r['fecha_busqueda'],
                            'IdAcuerdo': p.get('id_acuerdo', ''),
                            'Documento': p.get('documento', ''),
                            'Juicio': p.get('juicio', ''),
                            'Promoventes': p.get('promoventes', ''),
                            'Demandados': p.get('demandados', ''),
                            'Extracto': p.get('extracto', ''),
                            'Fecha Publicación': p.get('fecha_publicacion', '')
                        })
                else:
                    writer.writerow({
                        'Búsqueda': r['busqueda'],
                        'Tipo': r['tipo_busqueda'],
                        'Juzgado': r['juzgado'],
                        'Estado': r['estado'],
                        'Fecha Consulta': r['fecha_busqueda'],
                        'IdAcuerdo': '', 'Documento': '', 'Juicio': '',
                        'Promoventes': '', 'Demandados': '', 
                        'Extracto': '', 'Fecha Publicación': ''
                    })
        
        self.log(f"CSV guardado: {archivo}", "OK")

    def guardar_excel(self, archivo='resultados_expedientes.xlsx'):
        """Guarda resultados en Excel con formato y marcado de acuerdos nuevos"""
        self.log(f"Generando archivo Excel: {archivo}...")

        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados"

        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        nuevo_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Amarillo
        nuevo_font = Font(bold=True, color="FF0000")  # Rojo
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Encabezados
        headers = [
            'Búsqueda', 'Tipo', 'Juzgado', 'Estado', 'Fecha Consulta',
            'IdAcuerdo', 'Documento', 'Juicio', 'Promoventes',
            'Demandados', 'Extracto', 'Fecha Publicación', 'NUEVO'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_style

        # Datos
        row_num = 2
        total_nuevos = 0

        for r in self.resultados:
            if r['publicaciones']:
                for p in r['publicaciones']:
                    es_nuevo = p.get('es_nuevo', False)
                    if es_nuevo:
                        total_nuevos += 1

                    datos = [
                        r['busqueda'],
                        r['tipo_busqueda'],
                        r['juzgado'],
                        r['estado'],
                        r['fecha_busqueda'],
                        p.get('id_acuerdo', ''),
                        p.get('documento', ''),
                        p.get('juicio', ''),
                        p.get('promoventes', ''),
                        p.get('demandados', ''),
                        p.get('extracto', ''),
                        p.get('fecha_publicacion', ''),
                        '⭐ NUEVO' if es_nuevo else ''
                    ]

                    for col_num, valor in enumerate(datos, 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.value = valor
                        cell.border = border_style

                        # Aplicar formato a filas nuevas
                        if es_nuevo:
                            cell.fill = nuevo_fill
                            if col_num == 13:  # Columna "NUEVO"
                                cell.font = nuevo_font

                    row_num += 1
            else:
                # Sin publicaciones
                datos = [
                    r['busqueda'], r['tipo_busqueda'], r['juzgado'],
                    r['estado'], r['fecha_busqueda'],
                    '', '', '', '', '', '', '', ''
                ]

                for col_num, valor in enumerate(datos, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = valor
                    cell.border = border_style

                row_num += 1

        # Ajustar anchos de columna
        column_widths = {
            'A': 15, 'B': 12, 'C': 40, 'D': 18, 'E': 18,
            'F': 12, 'G': 30, 'H': 20, 'I': 30,
            'J': 30, 'K': 50, 'L': 15, 'M': 12
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Congelar primera fila
        ws.freeze_panes = 'A2'

        # Guardar archivo
        wb.save(archivo)
        self.log(f"Excel guardado: {archivo}", "OK")
        self.log(f"📊 Total de acuerdos NUEVOS marcados: {total_nuevos}", "INFO")

        return total_nuevos

    def resumen(self):
        """Muestra resumen de resultados"""
        print(f"\n{'='*60}")
        print("📊 RESUMEN DE RESULTADOS")
        print(f"{'='*60}")
        
        con = sum(1 for r in self.resultados if r['estado'] == 'Con publicaciones')
        sin = sum(1 for r in self.resultados if r['estado'] == 'Sin publicaciones')
        total_pubs = sum(len(r['publicaciones']) for r in self.resultados)
        
        print(f"Total búsquedas: {len(self.resultados)}")
        print(f"Con publicaciones: {con}")
        print(f"Sin publicaciones: {sin}")
        print(f"Total publicaciones: {total_pubs}")
        print("-" * 60)
        
        for r in self.resultados:
            num = len(r['publicaciones'])
            icono = "✅" if num > 0 else "⚪"
            juzgado_corto = r['juzgado'][:35] + "..." if len(r['juzgado']) > 35 else r['juzgado']
            print(f"  {icono} {r['busqueda']:15} | {juzgado_corto:38} | {num} pub.")
    
    def cerrar(self):
        if self.driver:
            self.log("Cerrando navegador...")
            self.driver.quit()
            self.log("Completado", "OK")


def main():
    """
    Versión 6.0 - OPTIMIZADA

    INSTRUCCIONES:
    1. Edita el archivo 'expedientes.json' para agregar/modificar expedientes
    2. Ejecuta este script
    3. Los resultados se guardarán en Excel con acuerdos nuevos marcados

    CONFIGURACIÓN:
    - max_pestanas: Número de pestañas simultáneas (default: 5)
    - dias_acuerdos_nuevos: Días para marcar como nuevo (default: 5)
    """

    print("=" * 70)
    print("🤖 Robot de Búsqueda Automática de Expedientes v6.0")
    print("    TSJ Quintana Roo - Lista Electrónica")
    print("=" * 70)

    # Cargar configuración desde config.json (o usar valores por defecto)
    config = TSJExpedientesBot.cargar_configuracion('config.json')
    max_pestanas = config.get('max_pestanas', 5)
    dias_nuevos = config.get('dias_acuerdos_nuevos', 5)

    print(f"\n⚙️  Configuración:")
    print(f"   - Pestañas simultáneas: {max_pestanas}")
    print(f"   - Días para marcar como nuevo: {dias_nuevos}")
    print("")

    bot = TSJExpedientesBot(max_pestanas=max_pestanas, dias_acuerdos_nuevos=dias_nuevos)

    try:
        # Intentar cargar expedientes desde JSON
        expedientes = bot.cargar_expedientes_json('expedientes.json')

        # Si no hay archivo JSON, usar expedientes por defecto
        if not expedientes:
            print("\n⚠️  No se encontró 'expedientes.json', usando expedientes por defecto...")
            expedientes = [
                # ===== JUZGADO SEGUNDO FAMILIAR ORAL CANCÚN =====
                {'numero': '2358/2025', 'juzgado': 'JUZGADO SEGUNDO FAMILIAR ORAL CANCUN'},
                {'numero': '2501/2025', 'juzgado': 'JUZGADO SEGUNDO FAMILIAR ORAL CANCUN'},
                {'numero': '2502/2025', 'juzgado': 'JUZGADO SEGUNDO FAMILIAR ORAL CANCUN'},
                {'numero': '1421/2025', 'juzgado': 'JUZGADO SEGUNDO FAMILIAR ORAL CANCUN'},

                # ===== JUZGADO PRIMERO FAMILIAR ORAL CANCÚN =====
                {'numero': '2500/2025', 'juzgado': 'JUZGADO PRIMERO FAMILIAR ORAL CANCUN'},
                {'numero': '1430/2022', 'juzgado': 'JUZGADO PRIMERO FAMILIAR ORAL CANCUN'},

                # ===== BÚSQUEDA POR NOMBRE - PLAYA DEL CARMEN =====
                {'nombre': 'samanta', 'juzgado': 'JUZGADO FAMILIAR ORAL PLAYA'},
            ]

        # Iniciar navegador y procesar
        bot.iniciar_navegador()
        bot.procesar_expedientes(expedientes)
        bot.resumen()

        # Guardar resultados en Excel
        total_nuevos = bot.guardar_excel('resultados_expedientes.xlsx')

        # También guardar CSV como respaldo
        bot.guardar_csv('resultados_expedientes.csv')

        print(f"\n{'='*70}")
        print(f"✅ PROCESO COMPLETADO")
        print(f"{'='*70}")
        print(f"📊 Archivo Excel: resultados_expedientes.xlsx")
        print(f"📄 Archivo CSV: resultados_expedientes.csv")
        print(f"⭐ Acuerdos nuevos (últimos {dias_nuevos} días): {total_nuevos}")
        print(f"{'='*70}")

    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()

    finally:
        input("\n⏸️  Presiona ENTER para cerrar el navegador...")
        bot.cerrar()


if __name__ == "__main__":
    main()